"""
TikTok Ads GMV Max Creative Exporter — 工具函数
=================================================
配置加载、日期处理、路径解析（含 ~ 扩展）、xlsx → JSON 转换等。
"""

from __future__ import annotations

import json
import os
import re
import warnings
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any


# ── 配置 ─────────────────────────────────────────────────────────────────


def load_config() -> dict[str, Any]:
    """加载项目根目录下的 config.json。

    Returns:
        配置字典。

    Raises:
        FileNotFoundError: config.json 不存在。
    """
    config_path = _project_root() / "config.json"
    if not config_path.exists():
        example_path = _project_root() / "config.example.json"
        raise FileNotFoundError(
            f"找不到 config.json。先把 {example_path} 复制为 {config_path}。"
        )
    return json.loads(config_path.read_text(encoding="utf-8"))


# ── 路径 ─────────────────────────────────────────────────────────────────


def _project_root() -> Path:
    """返回项目根目录（该文件所在目录）。"""
    return Path(__file__).parent.resolve()


def expand_home(value: str) -> str:
    """展开路径开头的 ~ 为当前用户 home 目录。

    Args:
        value: 可能以 ~ 开头的路径字符串。

    Returns:
        展开后的路径字符串。
    """
    if not isinstance(value, str):
        return value
    if value == "~":
        return os.path.expanduser("~")
    if value.startswith("~/"):
        return os.path.join(os.path.expanduser("~"), value[2:])
    return value


def resolve_from_project(value: str | None) -> Path:
    """将路径解析为绝对路径。

    支持:
    - 空值/None → 项目根目录
    - ~/xxx → 用户 home 目录
    - 绝对路径 → 原样返回
    - 相对路径 → 相对于项目根目录

    Args:
        value: 路径字符串。

    Returns:
        绝对 Path 对象。
    """
    if not value:
        return _project_root()
    expanded = expand_home(value)
    p = Path(expanded)
    return p if p.is_absolute() else _project_root() / p


def ensure_dir(dir_path: str | Path) -> None:
    """确保目录存在（类似 mkdir -p）。"""
    Path(dir_path).mkdir(parents=True, exist_ok=True)


# ── 日期 ─────────────────────────────────────────────────────────────────


def yesterday() -> datetime:
    """返回昨天的 00:00:00 日期时间对象。"""
    dt = datetime.now() - timedelta(days=1)
    return dt.replace(hour=0, minute=0, second=0, microsecond=0)


def format_date(dt: datetime, fmt: str = "yyyy-MM-dd") -> str:
    """按指定格式格式化日期。

    支持的占位符: yyyy, MM, dd

    Args:
        dt: 日期时间对象。
        fmt: 格式字符串。

    Returns:
        格式化后的日期字符串。
    """
    return (
        fmt.replace("yyyy", str(dt.year))
        .replace("MM", f"{dt.month:02d}")
        .replace("dd", f"{dt.day:02d}")
    )


def safe_filename_part(value: str, max_length: int = 90) -> str:
    """将字符串清理为安全的文件名片段。

    替换非法字符，合并空白，截断到指定长度。

    Args:
        value: 原始字符串。
        max_length: 最大长度（默认 90）。

    Returns:
        安全的文件名片段。
    """
    cleaned = re.sub(r'[\\/:*?"<>|]', "-", str(value))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned[:max_length]


def timestamp_for_filename(dt: datetime | None = None) -> str:
    """返回紧凑时间戳（HHmmss）。

    Args:
        dt: 日期时间对象，默认当前时间。

    Returns:
        六位时间字符串。
    """
    if dt is None:
        dt = datetime.now()
    return f"{dt.hour:02d}{dt.minute:02d}{dt.second:02d}"


# ── xlsx → JSON 转换 ────────────────────────────────────────────────────


def xlsx_to_json(xlsx_path: str | Path) -> dict[str, list[dict[str, Any]]]:
    """将 xlsx 文件转换为 JSON 数据结构。

    每个 sheet 为一个 key，value 为 {header: value} 字典列表。
    第一行为表头，空行跳过，datetime 转为 ISO 字符串。

    Args:
        xlsx_path: .xlsx 文件路径。

    Returns:
        {sheet_name: [{col: val, ...}, ...], ...}
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"文件不存在: {xlsx_path}")

    try:
        import openpyxl
    except ImportError:
        raise ValueError("需要 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            result[sheet_name] = []
            continue

        headers = [_clean_header(h, i) for i, h in enumerate(rows[0])]
        data: list[dict[str, Any]] = []

        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            row_dict: dict[str, Any] = {}
            for i, val in enumerate(row):
                if i >= len(headers):
                    break
                row_dict[headers[i]] = _serialize_value(val)
            data.append(row_dict)

        result[sheet_name] = data

    wb.close()
    return result


def _clean_header(header: Any, index: int) -> str:
    """清理表头，空值使用 fallback 名称。"""
    if header is None or str(header).strip() == "":
        return f"column_{index}"
    return str(header).strip()


def _serialize_value(val: Any) -> Any:
    """将单元格值转为 JSON 兼容类型。"""
    if val is None:
        return None
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return val


def save_json_output(
    data: dict[str, list[dict[str, Any]]],
    json_path: str | Path,
    ensure_ascii: bool = False,
    indent: int = 2,
) -> Path:
    """将数据保存为 JSON 文件。

    Args:
        data: 数据。
        json_path: 输出路径。
        ensure_ascii: 是否转义非 ASCII（默认 False，保留中文）。
        indent: 缩进空格数。

    Returns:
        JSON 文件路径。
    """
    json_path = Path(json_path)
    json_path.write_text(
        json.dumps(data, ensure_ascii=ensure_ascii, indent=indent, default=str),
        encoding="utf-8",
    )
    return json_path


def convert_xlsx_in_dir(xlsx_path: str | Path) -> Path | None:
    """将 xlsx 文件转换为同名的 .json 文件（如果 json 已更新则跳过）。"""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists() or xlsx_path.suffix.lower() not in (".xlsx", ".xls"):
        return None

    json_path = xlsx_path.with_suffix(".json")
    if json_path.exists() and json_path.stat().st_mtime >= xlsx_path.stat().st_mtime:
        return json_path

    try:
        data = xlsx_to_json(xlsx_path)
        save_json_output(data, json_path)
        return json_path
    except Exception as exc:
        warnings.warn(f"转换 {xlsx_path.name} 到 JSON 失败: {exc}")
        return None
